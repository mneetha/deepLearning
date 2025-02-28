import pdfplumber
import pandas as pd
import re

from openpyxl.workbook import Workbook
from textblob import TextBlob
import textdistance
from symspellpy import SymSpell, Verbosity
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
from openpyxl import load_workbook
import os
import glob

# Path to your PDF
pdf_paths = glob.glob("../../Desktop/financials/*.pdf")[:1]


def extract_income_expense_data(text):
    category_pattern = r"^(?!Total )([A-Za-z &:/]+(?: Expenses| Services|Utilities| Contract| Reserve| Operating| Income)?)$"
    expense_pattern = r'(\d{5})\s-\s([A-Za-z &:/\-,]+)((?:\s*([\d,]+\.\d{2}|\-))+)'
    # Data storage
    expenses_by_category = {}
    current_category = None
    lines = text.split("\n")

    for line in lines:
        line = line.strip()

        # Check for category
        category_match = re.match(category_pattern, line)
        if category_match:
            current_category = category_match.group(1)
            if current_category not in expenses_by_category:
                expenses_by_category[current_category] = []
            continue  # Move to next line

        # Check for expense entry
        expense_match = re.search(expense_pattern, line)
        if expense_match and current_category:
            code = expense_match.group(1)
            title = expense_match.group(2)
            cost = expense_match.group(3)
            # replace - with 0
            match_hyphen = re.search(r"(\s*-+\s*)+$", title)
            if match_hyphen:
                # print(f'found {title}@{cost}')
                match_hyphen_str = match_hyphen.group()
                title = re.sub(r"(\s*-+\s*)+$", '', title)
                # print(f'edited {title}@{cost}')
            else:
                match_hyphen_str = ''

            new_cost= match_hyphen_str + cost
            new_cost_str = new_cost.replace('-', '0')
            expenses_by_category[current_category].append({
                "code": code,
                "title": title,
                "cost_incurred": new_cost_str.strip().split()[0]
            })

    return expenses_by_category

def extract_balance_sheet_data(text):
    category_pattern = r"^(?!Total )(?:Cash-Operating|Cash-Reserve|Accounts Receivable|Other Assets|Accounts Payable|Other Liabilities|Fund Balance)$"
    balance_pattern = r'(\d{5})\s-\s([A-Za-z &:/,\d{1,2}\-\%\.]+)\s(\$\d{1,3}(?:,\d{3})*\.\d{2})(\s\$\d{1,3}(?:,\d{3})*\.\d{2})*'
    balances_by_category = {}
    current_category = None
    lines = text.split("\n")

    for line in lines:
        line = line.strip()
        # Check for category
        category_match = re.match(category_pattern, line)
        if category_match:
            current_category = category_match.group()
            balances_by_category[current_category] = []
            continue  # Move to next line

        # Check for expense entry
        account_match = re.search(balance_pattern, line)
        if account_match and current_category:
            code = account_match.group(1)
            acct = account_match.group(2)
            balance = account_match.group(3)
            balances_by_category[current_category].append({
                "code": code,
                "acct": acct,
                "balance": balance
            })

    # print(balances_by_category)
    return balances_by_category

def extract_ar_aging_data(text):
    # Regex pattern to extract unit details and overdue amounts
    # Unit Details Pattern
    unit_pattern = re.compile(
        r"(?P<unit_id>\d{9}) - (?P<address>[\d\s\w]+) Unit (?P<unit_number>[\w\d-]+) - (?P<resident>[\w\s,]+)"
    )

    payment_pattern = re.compile(
        r'(?P<category>Monthly Assessment|Late Fee|Violation|Reimbursement Assessment|Collection Charges| Charges|Maintenance & Repair)\s+\$?(?P<due_0_30>[\d,]+\.\d{2}|-)\s+\$?(?P<due_30_60>[\d,]+\.\d{2}|-)\s+\$?(?P<due_60_90>[\d,]+\.\d{2}|-)\s+\$?(?P<due_over_90>[\d,]+\.\d{2}|-)\s+\$(?P<total_due>[\d,]+\.\d{2})'
    )

    lines = text.split("\n")
    ar_aging = {}
    for line in lines:
        line = line.strip()
        unit_match = re.search(unit_pattern, line)
        if not unit_match:
            payment_match = re.search(payment_pattern,line)
            if payment_match:
                if current_unit_id:
                    payment_dict = {}
                    payment_dict["category"] = payment_match.group("category")
                    payment_dict["due_30"] = payment_match.group("due_0_30").replace('-','0')
                    payment_dict["due_60"] = payment_match.group("due_30_60").replace('-','0')
                    payment_dict["due_90"] = payment_match.group("due_60_90").replace('-','0')
                    payment_dict["over_90"] = payment_match.group("due_over_90").replace('-','0')
                    payment_dict["total_due"] = payment_match.group("total_due").replace('-','0')
                    ar_aging[current_unit_id].append(payment_dict)
        else:
            current_unit_id = unit_match.group("unit_id")
            ar_aging[current_unit_id] = []

    print(ar_aging)
    return ar_aging

def extract_cash_disburse_data(text):
    # print(text)
    cd_1 = re.compile(
        r"(?P<date_>\d{1,2}/\d{1,2}/\d{4})\s+"  # Date (e.g., 11/4/2024)
        r"(?P<chk_num>(Avid\s+\d+|Transfer\s+Out|ACH\s+[\w\s]+))\s+"  # Check number (e.g., Avid 100280 or Transfer Out)
        r"(?P<desc>.+?)\s+"  # Description (allowing for multi-word descriptions with spaces)
        r"(?P<amt>[\d,]+\.\d{2})\s*$"  # Amount (e.g., 147.38)
    )

    cd_2 = re.compile(
        r"(?P<code>\d+)\s+-\s+"                  # Capturing reference ID (e.g., 55555 or 10100)
        r"(?P<desc>.+?)\s+"          # Capturing description of work/transfer details
        r"(?P<sub_amt>[\d,]+\.\d{2})"              # Capturing amount (e.g., 111.38, 5,170.00)
    )

    cd_ = {}
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        cd_1_match = re.search(cd_1, line)
        if not cd_1_match:
            cd_2_match = re.search(cd_2, line)
            if cd_2_match:
                if current_chk_num:
                    code_breakdown = {}
                    code_breakdown['code'] = cd_2_match.group('code')
                    code_breakdown['sub_amt'] = cd_2_match.group('sub_amt')
                    cd_[current_chk_num].append(code_breakdown)

        else:
            current_chk_num = cd_1_match.group('chk_num')
            if current_chk_num in cd_:
                continue
            else:
                cd_[current_chk_num] = []
    #print(cd_)
    return cd_

def extract_bank_reconcile_data(text):
    print(text)
    lines = text.split("\n")
    
        # get reconciliation summary

def filter_financial_values_from_pdf(pdf_paths):
    #create the excel file..
    file_name = 'rosewalk_financials.xlsx'
    for pdf_path in pdf_paths:
        print(f"Working on file {pdf_path}")
        with pdfplumber.open(pdf_path) as pdf:

            ## extract balance sheet data
            BALANCE_SHEET = 'Balance Sheet'
            balance_sheet_text = ''
            balance_sheet_regex = r'(Balance Sheet) as of (\d{1,2}/\d{1,2}/\d{4})'

            INCOME_EXPENSE = 'Statement of Revenues and Expenses'
            income_expense_text = ''
            income_expense_regex = r'(Statement of Revenues and Expenses)'

            AP_AGING = 'AP Aging for Ending Date'
            ap_aging_text = ''
            ap_aging_regex = r'(AP Aging for Ending Date)'

            AR_AGING = 'AR Aging'
            ar_aging_text = ''
            ar_aging_regex = r'(AR Aging)'

            CASH_DISBURSE = 'Cash Disbursement'
            cd_text = ''
            cd_regex = r'(Cash Disbursement)'

            BANK_RECONCILE = 'Bank Account Reconciliation'
            br_text = ''
            br_regex = r'(Bank Account Reconciliation)'


            total_pages = len(pdf.pages)
            for i, page in enumerate(pdf.pages, start=2):
                if i - 1 >= total_pages:  # Prevent index out of range
                    break
                current_page = i
                page = pdf.pages[i-1]
                page_text = page.extract_text()
                match_bal = re.search(balance_sheet_regex, page_text)
                if match_bal and match_bal.group(1) == BALANCE_SHEET:
                    extracted_date = match_bal.group(2)
                    balance_sheet_text += page_text
                match_exp = re.search(income_expense_regex, page_text)
                if match_exp and match_exp.group(1) == INCOME_EXPENSE:
                    income_expense_text += page_text
                match_ap = re.search(ap_aging_regex, page_text)
                if match_ap and match_ap.group(1) == AP_AGING:
                    ap_aging_text += page_text
                match_ar = re.search(ar_aging_regex, page_text)
                if match_ar and match_ar.group(1) == AR_AGING:
                    ar_aging_text += page_text
                match_cd = re.search(cd_regex, page_text)
                if match_cd and match_cd.group(1) == CASH_DISBURSE:
                    cd_text += page_text
                match_br = re.search(br_regex, page_text)
                if match_br and match_br.group(1) == BANK_RECONCILE:
                    br_text += page_text

            # bal_data_to_write = extract_balance_sheet_data(balance_sheet_text)
            # income_expense_data_to_write = extract_income_expense_data(income_expense_text)
            # ap_aging_data_to_write = extract_ap_aging_data(ap_aging_text)
            # ar_aging_data_to_write = extract_ar_aging_data(ar_aging_text)
            # cash_disburse_to_write = extract_cash_disburse_data(cd_text)
            bank_reconcile_to_write = extract_bank_reconcile_data(br_text)
            '''
            sheet_name = extracted_date.replace("/", "-")

            if os.path.exists(file_name):
                book = load_workbook(file_name)
                if sheet_name in book.sheetnames:
                    del book[sheet_name]
                sheet = book.create_sheet(sheet_name)
            else:
                book = Workbook()
                sheet = book.active
                sheet.title = sheet_name
            book.save(file_name)

            startrow = sheet.max_row + 2

            with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                writer._book = book  # Attach existing workbook

                # Save workbook before writing data
                book.save(file_name)

                # Function to write a dictionary to Excel
                def write_dict_to_excel(data_dict, section_title, section_header, start_row, columns_to_write):
                    """Writes dictionary data with a section title into Excel."""
                    sheet[f"A{start_row}"] = section_title  # Insert section title
                    start_row += 1  # Move to next row

                    for category, entries in data_dict.items():
                        if not entries:
                            continue  # Skip empty categories

                        df = pd.DataFrame(entries, columns=columns_to_write)
                        df.insert(0, section_header, category)  # Insert category in the first column

                        # Write to Excel
                        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False,
                                    header=(start_row == 2))
                        start_row += len(df) + 2  # Leave space before next section

                    return start_row  # Return updated row position

                print(f'on row# {startrow}')
                # Write first dictionary with "Balance Sheet" title
                startrow = write_dict_to_excel(bal_data_to_write, "Balance Sheet", "Account Type", startrow,
                                               columns_to_write=["code", "acct", "balance"])
                print(f'on row# {startrow}')
                # Write second dictionary with "Income & Expenses" title
                startrow = write_dict_to_excel(income_expense_data_to_write, "Income & Expenses", "Income/Expense type", startrow,
                                               columns_to_write=["code", "title", "cost_incurred"])
                print(f'on row# {startrow}')
                # Write third dictionary with "AR Aging" title
                startrow = write_dict_to_excel(ar_aging_data_to_write, "AR Aging", "Unit ID", startrow,
                                               columns_to_write=["category", "due_30", "due_60", "due_90", "over_90", "total_due"])
                print(f'on row# {startrow}')
                
                print(f'on row# {startrow}')
                # Write third dictionary with "Cash Disbursement" title
                startrow = write_dict_to_excel(cash_disburse_to_write, "Cash Disbursement", "chk_num", startrow,
                                               columns_to_write=["code", "sub_amt"])
                print(f'on row# {startrow}')
            # Save workbook
            writer.book.save(file_name)

        '''
filter_financial_values_from_pdf(pdf_paths)
